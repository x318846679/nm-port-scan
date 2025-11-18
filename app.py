from functools import wraps
from flask import Flask, render_template, redirect, url_for, session, request, jsonify, flash
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SubmitField, TextAreaField, SelectField, IntegerField, BooleanField
from wtforms.validators import DataRequired, Length, Email, NumberRange
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import subprocess
import json
import nmap
import ipaddress
import sqlite3
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# 创建线程池执行器
executor = ThreadPoolExecutor(max_workers=20)

terminate_flags = {} # 用于存储每个任务的终止标志

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///nm_scan.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = '请先登录以访问此页面。'


# Forms
class LoginForm(FlaskForm):
    username = StringField('用户名', validators=[DataRequired()])
    password = PasswordField('密码', validators=[DataRequired()])
    submit = SubmitField('登录')


class ChangePasswordForm(FlaskForm):
    old_password = PasswordField('旧密码', validators=[DataRequired()])
    new_password = PasswordField('新密码', validators=[DataRequired(), Length(min=8)])
    confirm_password = PasswordField('确认新密码', validators=[DataRequired()])
    submit = SubmitField('修改密码')


class UserForm(FlaskForm):
    username = StringField('用户名', validators=[DataRequired()])
    password = PasswordField('密码', validators=[DataRequired(), Length(min=8)])
    confirm = PasswordField('确认密码', validators=[DataRequired()])
    role = SelectField('角色', choices=[('user', '普通用户'), ('admin', '管理员')], validators=[DataRequired()])
    email = StringField('邮箱')  # 移除Email()验证器，使其成为可选字段
    submit = SubmitField('保存')
    
    def validate_confirm(self, field):
        if self.password.data != field.data:
            raise ValidationError('密码和确认密码不匹配')


class TemplateForm(FlaskForm):
    name = StringField('模板名称', validators=[DataRequired()])
    params = TextAreaField('参数', validators=[DataRequired()], render_kw={"rows": 10})
    description = TextAreaField('描述', render_kw={"rows": 3})
    submit = SubmitField('保存')


class AssetForm(FlaskForm):
    name = StringField('资产名称', validators=[DataRequired()])
    ip = TextAreaField('IP地址', validators=[DataRequired()], render_kw={"rows": 10})
    submit = SubmitField('保存')


class TaskForm(FlaskForm):
    name = StringField('任务名称', validators=[DataRequired()])
    asset_id = SelectField('选择资产', coerce=int, validators=[])
    template_id = SelectField('选择模板', coerce=int, validators=[])
    custom_ips = TextAreaField('手动输入IP列表 (每行一个)', render_kw={"rows": 5})
    send_email = BooleanField('扫描完成后发送结果到邮箱')
    email_address = StringField('邮箱地址')
    schedule_type = SelectField('执行方式', choices=[
        ('once', '一次性'),
        ('daily', '每天'),
        ('weekly', '每周'),
        ('monthly', '每月')
    ], default='once')
    submit = SubmitField('开始扫描')
    
    def validate(self, extra_validators=None):
        # 调用父类的validate方法
        if not super(TaskForm, self).validate(extra_validators=extra_validators):
            return False
            
        # 额外的自定义验证
        # 检查是否提供了IP地址（通过资产或自定义输入）
        if (not self.custom_ips.data or not self.custom_ips.data.strip()) and \
           (not self.asset_id.data or self.asset_id.data == -1):
            self.custom_ips.errors.append('请提供要扫描的IP地址，可以通过选择资产或手动输入')
            return False
            
        # 如果选择了发送邮件，检查邮箱地址
        if self.send_email.data and (not self.email_address.data or not self.email_address.data.strip()):
            self.email_address.errors.append('请选择发送邮件选项时，必须提供邮箱地址')
            return False
            
        # 如果提供了邮箱地址，验证邮箱格式
        if self.email_address.data and self.email_address.data.strip():
            email_validator = Email()
            try:
                email_validator(self, self.email_address)
            except ValidationError:
                self.email_address.errors.append('邮箱地址格式不正确')
                return False
                
        return True


class EmailForm(FlaskForm):
    host = StringField('SMTP服务器', validators=[DataRequired()])
    port = IntegerField('端口', validators=[DataRequired(), NumberRange(1, 65535)])
    user = StringField('用户名', validators=[DataRequired()])
    passwd = PasswordField('密码', validators=[DataRequired()])
    submit = SubmitField('保存')


# Models
class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(80), nullable=False, default='user')
    # 将email字段标记为可选，以兼容已存在的数据库结构
    email = db.Column(db.String(120), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    def __repr__(self):
        return f'<User {self.username}>'


class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    ips = db.Column(db.Text, nullable=False)
    params = db.Column(db.Text, nullable=True)
    status = db.Column(db.String(20), default='pending')  # pending, running, completed, failed
    result = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    finished_at = db.Column(db.DateTime, nullable=True)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    progress = db.Column(db.Integer, default=0)
    send_email = db.Column(db.Boolean, default=False)  # 是否发送邮件
    email_address = db.Column(db.String(120), nullable=True)  # 接收邮件的地址
    schedule_type = db.Column(db.String(20), default='once')  # once, daily, weekly, monthly
    last_run_at = db.Column(db.DateTime, nullable=True)  # 上次执行时间
    
    def __repr__(self):
        return f'<Task {self.name}>'


class Asset(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ip = db.Column(db.String(50), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    def __repr__(self):
        return f'<Asset {self.name}>'


class ScanTemplate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    params = db.Column(db.Text, nullable=False)
    description = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    def __repr__(self):
        return f'<ScanTemplate {self.name}>'


class ZEmail(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    host = db.Column(db.String(100), nullable=False)
    port = db.Column(db.Integer, nullable=False)
    user = db.Column(db.String(100), nullable=False)
    passwd = db.Column(db.String(100), nullable=False)
    
    def __repr__(self):
        return f'<ZEmail {self.host}>'


class AccessLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), nullable=False)
    ip_address = db.Column(db.String(45), nullable=False)
    user_agent = db.Column(db.Text, nullable=True)
    login_time = db.Column(db.DateTime, default=datetime.now)
    failed_login_attempts = db.Column(db.Integer, default=0)
    block_until = db.Column(db.DateTime, nullable=True)
    request_url = db.Column(db.String(200), nullable=True)
    
    def __repr__(self):
        return f'<AccessLog {self.username}>'

# FOFA搜索结果模型
class FofaResult(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ip = db.Column(db.String(45), nullable=False)
    port = db.Column(db.Integer, nullable=False)
    protocol = db.Column(db.String(20), nullable=True)
    domain = db.Column(db.String(255), nullable=True)
    title = db.Column(db.Text, nullable=True)
    server = db.Column(db.String(255), nullable=True)
    os = db.Column(db.String(100), nullable=True)
    country = db.Column(db.String(10), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    def to_dict(self):
        return {
            'id': self.id,
            'ip': self.ip,
            'port': self.port,
            'protocol': self.protocol,
            'domain': self.domain,
            'title': self.title,
            'server': self.server,
            'os': self.os,
            'country': self.country,
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M:%S') if self.created_at else None
        }


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def log_access(username):
    """记录用户访问日志"""
    ip_address = request.remote_addr
    access_log = AccessLog(
        username=username,
        ip_address=ip_address,
        request_url=request.path
    )
    db.session.add(access_log)
    db.session.commit()


# Routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and check_password_hash(user.password, form.password.data):
            login_user(user, remember=False)
            # 确保在session中设置用户角色
            session['role'] = user.role
            session['username'] = user.username
            log_access(user.username)  # 记录成功登录
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('index'))
        else:
            ip_address = request.remote_addr
            access_log = AccessLog.query.filter_by(username=form.username.data, ip_address=ip_address).first()
            if access_log:
                access_log.failed_login_attempts += 1
                if access_log.failed_login_attempts > 10:
                    access_log.block_until = datetime.now() + timedelta(minutes=10)  # 封禁10分钟
            else:
                access_log = AccessLog(username=form.username.data, ip_address=ip_address, failed_login_attempts=1, request_url=request.path)
                db.session.add(access_log)
            db.session.commit()
            
            if access_log.failed_login_attempts > 10:
                return redirect(url_for('attack_warning'))  # 跳转到警告页面
                
    return render_template('login.html', form=form)


@app.route('/attack_warning')
def attack_warning():
    return render_template('attack_warning.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/', methods=['GET', 'POST'])
@login_required
def index():
    print("进入index路由")
    form = TaskForm()
    print("表单实例化完成")
    
    # Populate choices for assets and templates
    form.asset_id.choices = [(-1, '不选择')] + [(asset.id, asset.name) for asset in Asset.query.all()]
    form.template_id.choices = [(-1, '不选择')] + [(template.id, template.name) for template in ScanTemplate.query.all()]
    print("资产和模板选项已加载")
    
    print(f"请求方法: {request.method}")
    if request.method == 'POST':
        print("收到POST请求")
        print("表单数据:")
        for key, value in request.form.items():
            print(f"  {key}: {value}")
    
    if form.validate_on_submit():
        print("表单验证通过")
        try:
            print("开始处理任务创建")
            print("表单数据:")
            print(f"  任务名称: {form.name.data}")
            print(f"  资产ID: {form.asset_id.data}")
            print(f"  模板ID: {form.template_id.data}")
            print(f"  自定义IP: {form.custom_ips.data}")
            print(f"  发送邮件: {form.send_email.data}")
            print(f"  邮箱地址: {form.email_address.data}")
            print(f"  调度类型: {form.schedule_type.data}")
            
            # Determine IPs to scan
            ips = ""
            if form.custom_ips.data and form.custom_ips.data.strip():
                ips = form.custom_ips.data.strip()
                print(f"使用自定义IP: {ips}")
            elif form.asset_id.data is not None and form.asset_id.data != -1:
                print(f"尝试获取资产ID: {form.asset_id.data}")
                asset = Asset.query.get(form.asset_id.data)
                if asset:
                    ips = asset.ip
                    print(f"使用资产IP: {ips}")
                else:
                    print("未能获取资产信息")
            else:
                print("未提供IP地址")
            
            # Check if we have IPs to scan
            if not ips:
                flash('请提供要扫描的IP地址，可以通过选择资产或手动输入', 'error')
                print("错误: 未提供IP地址")
            else:
                # Get template parameters or custom parameters
                params = ""
                # Check for custom parameters first
                custom_params = request.form.get('custom_params', '').strip()
                if custom_params:
                    params = custom_params
                    print(f"使用手动输入参数: {params}")
                elif form.template_id.data is not None and form.template_id.data != -1:
                    print(f"尝试获取模板ID: {form.template_id.data}")
                    template = ScanTemplate.query.get(form.template_id.data)
                    if template:
                        params = template.params
                        print(f"使用模板参数: {params}")
                    else:
                        print("未能获取模板参数")
                
                # Create task
                print("开始创建任务...")
                task = Task(
                    name=form.name.data,
                    ips=ips,
                    params=params,
                    created_by=current_user.id,
                    send_email=form.send_email.data,
                    email_address=form.email_address.data if form.send_email.data else None,
                    schedule_type=form.schedule_type.data
                )
                db.session.add(task)
                db.session.flush()  # 获取任务ID但不提交
                task_id = task.id
                print(f"任务已创建，ID: {task_id}")
                
                db.session.commit()
                print("任务已提交到数据库")
                
                # Run scan in background
                print("提交任务到后台执行...")
                executor.submit(run_scan, task_id)
                flash('任务创建成功，正在后台扫描', 'success')
                print("任务创建成功")
                
        except Exception as e:
            db.session.rollback()
            flash(f'任务创建失败: {str(e)}', 'error')
            print(f"任务创建失败: {str(e)}")
            import traceback
            print(traceback.format_exc())
        
        print("重定向到首页")
        return redirect(url_for('index'))
    else:
        print("表单验证失败")
        if form.errors:
            print("表单错误:")
            for field, errors in form.errors.items():
                print(f"  {field}: {errors}")
                flash(f'字段 {field} 错误: {", ".join(errors)}', 'error')
        if request.method == 'POST':
            print("POST请求但表单验证失败")
            flash('表单验证失败，请检查输入信息', 'error')

    print("渲染首页模板")
    # Get statistics
    total_tasks = Task.query.count()
    running_tasks = Task.query.filter_by(status='running').count()
    completed_tasks = Task.query.filter_by(status='completed').count()
    failed_tasks = Task.query.filter_by(status='failed').count()
    pending_tasks = Task.query.filter_by(status='pending').count()
    total_assets = Asset.query.count()
    total_templates = ScanTemplate.query.count()
    
    # Calculate total ports
    total_ports = 0
    completed_tasks_data = Task.query.filter_by(status='completed').all()
    for task in completed_tasks_data:
        if task.result:
            try:
                ports = json.loads(task.result)
                for host_info in ports:
                    total_ports += len(host_info.get('ports', []))
            except:
                pass

    # Get recent tasks
    recent_tasks = Task.query.order_by(Task.created_at.desc()).limit(5).all()
    
    # Prepare chart data
    # Port distribution chart (top services)
    port_stats = {}
    for task in completed_tasks_data:
        if task.result:
            try:
                ports = json.loads(task.result)
                for port in ports:
                    service = port.get('service', 'unknown')
                    port_stats[service] = port_stats.get(service, 0) + 1
            except:
                pass
    
    # Top 5 services
    sorted_stats = sorted(port_stats.items(), key=lambda x: x[1], reverse=True)[:5]
    chart_labels = [item[0] for item in sorted_stats]
    chart_data = [item[1] for item in sorted_stats]
    
    # Weekly activity (last 7 days)
    from datetime import datetime, timedelta
    today = datetime.now().date()
    week_labels = []
    week_data = []
    
    for i in range(6, -1, -1):  # Last 7 days including today
        date = today - timedelta(days=i)
        count = Task.query.filter(
            db.func.date(Task.created_at) == date
        ).count()
        week_labels.append(date.strftime('%m-%d'))
        week_data.append(count)
    
    weekly_activity = {
        'labels': week_labels,
        'data': week_data
    }
    
    stats = {
        'total_tasks': total_tasks,
        'running_tasks': running_tasks,
        'completed_tasks': completed_tasks,
        'total_assets': total_assets,
        'total_templates': total_templates,
        'total_ports': total_ports
    }
    
    chart_data_obj = {
        'labels': chart_labels,
        'data': chart_data
    }
    
    return render_template('index.html', 
                         form=form, 
                         stats={
                             'total_tasks': total_tasks,
                             'running_tasks': running_tasks,
                             'completed_tasks': completed_tasks,
                             'failed_tasks': failed_tasks,
                             'pending_tasks': pending_tasks,
                             'total_assets': total_assets,
                             'total_ports': total_ports
                         },
                         recent_tasks=recent_tasks,
                         weekly_activity=weekly_activity)


@app.route('/tasks')
@login_required
def tasks_list():
    tasks = Task.query.all()
    return render_template('task_list.html', tasks=tasks)


@app.route('/task/<int:task_id>')
@login_required
def task_detail(task_id):
    task = Task.query.get_or_404(task_id)
    
    # 获取端口数据
    ports_data = []
    if task.result:
        try:
            ports_data = json.loads(task.result)
        except:
            ports_data = []

    # ports_data的格式为 [{'host': '153.3.238.127', 'hostname': 'www.baidu.com', 'state': 'up', 'ports': [{'port': 80, 'state': 'open', 'service': 'http', 'product': '', 'version': ''}]}]
    
    # 获取对比数据（与上一次扫描结果对比）
    diff_data = None
    # 查找相同资产（ips相同）的上一次完成的任务
    previous_task = Task.query.filter(
        Task.id < task_id,
        Task.ips == task.ips,
        Task.status == 'completed'
    ).order_by(Task.id.desc()).first()
    
    if previous_task and previous_task.result:
        try:
            # 解析当前和历史端口数据
            current_ports_set = set()
            previous_ports_set = set()
            
            # 构建当前端口集合
            for host_data in ports_data:
                host = host_data.get('host', '')
                if 'ports' in host_data and isinstance(host_data['ports'], list):
                    for port_data in host_data['ports']:
                        # 只考虑开放的端口
                        if port_data.get('state') == 'open':
                            key = f"{host}:{port_data.get('port', '')}:{port_data.get('protocol', 'tcp')}"
                            current_ports_set.add(key)
            
            # 构建历史端口集合
            previous_data = json.loads(previous_task.result)
            for host_data in previous_data:
                host = host_data.get('host', '')
                if 'ports' in host_data and isinstance(host_data['ports'], list):
                    for port_data in host_data['ports']:
                        # 只考虑开放的端口
                        if port_data.get('state') == 'open':
                            key = f"{host}:{port_data.get('port', '')}:{port_data.get('protocol', 'tcp')}"
                            previous_ports_set.add(key)
            
            # 计算差异
            added_keys = current_ports_set - previous_ports_set
            removed_keys = previous_ports_set - current_ports_set
            unchanged_keys = current_ports_set & previous_ports_set
            
            # 构建差异数据
            added_ports = []
            removed_ports = []
            unchanged_ports = []
            
            # 为新增端口构建详细信息
            for host_data in ports_data:
                host = host_data.get('host', '')
                if 'ports' in host_data and isinstance(host_data['ports'], list):
                    for port_data in host_data['ports']:
                        key = f"{host}:{port_data.get('port', '')}:{port_data.get('protocol', 'tcp')}"
                        if key in added_keys:
                            port_info = {
                                'host': host,
                                'port': port_data.get('port', 'N/A'),
                                'protocol': port_data.get('protocol', 'tcp'),
                                'service': port_data.get('service', 'unknown')
                            }
                            added_ports.append(port_info)
            
            # 为减少的端口构建详细信息
            previous_data = json.loads(previous_task.result)
            for host_data in previous_data:
                host = host_data.get('host', '')
                if 'ports' in host_data and isinstance(host_data['ports'], list):
                    for port_data in host_data['ports']:
                        key = f"{host}:{port_data.get('port', '')}:{port_data.get('protocol', 'tcp')}"
                        if key in removed_keys:
                            port_info = {
                                'host': host,
                                'port': port_data.get('port', 'N/A'),
                                'protocol': port_data.get('protocol', 'tcp'),
                                'service': port_data.get('service', 'unknown')
                            }
                            removed_ports.append(port_info)
            
            # 为未变的端口构建详细信息
            for host_data in ports_data:
                host = host_data.get('host', '')
                if 'ports' in host_data and isinstance(host_data['ports'], list):
                    for port_data in host_data['ports']:
                        key = f"{host}:{port_data.get('port', '')}:{port_data.get('protocol', 'tcp')}"
                        if key in unchanged_keys:
                            port_info = {
                                'host': host,
                                'port': port_data.get('port', 'N/A'),
                                'protocol': port_data.get('protocol', 'tcp'),
                                'service': port_data.get('service', 'unknown')
                            }
                            unchanged_ports.append(port_info)
            
            diff_data = {
                'previous_task': {
                    'id': previous_task.id,
                    'name': previous_task.name,
                    'created_at': previous_task.created_at.strftime('%Y-%m-%d %H:%M:%S')
                },
                'added_ports': added_ports,
                'removed_ports': removed_ports,
                'unchanged_ports': unchanged_ports
            }
        except Exception as e:
            print(f"处理对比数据时出错: {e}")
            # 即使处理对比数据出错，也要确保diff_data为一个空对象而不是None
            diff_data = {}
    
    # 获取历史数据（用于趋势图）
    history_data = []
    history_tasks = Task.query.filter(
        Task.ips == task.ips,
        Task.status == 'completed'
    ).order_by(Task.created_at).all()
    
    for history_task in history_tasks:
        if history_task.result:
            try:
                result_data = json.loads(history_task.result)
                port_count = 0
                # 计算总端口数（只计算开放的端口）
                for host_data in result_data:
                    if 'ports' in host_data and isinstance(host_data['ports'], list):
                        # 只计算开放的端口
                        open_ports = [port for port in host_data['ports'] if port.get('state') == 'open']
                        port_count += len(open_ports)
                
                history_data.append({
                    'task_id': history_task.id,
                    'created_at': history_task.created_at.strftime('%Y-%m-%d %H:%M'),
                    'port_count': port_count
                })
            except Exception as e:
                print(f"处理历史数据时出错: {e}")
                pass
    
    return render_template('task_detail.html', 
                         task=task, 
                         ports_data=ports_data,
                         diff_data=diff_data if diff_data is not None else {},
                         history_data=history_data)


@app.route('/assets', methods=['GET', 'POST'])
@login_required
def assets():
    form = AssetForm()
    asset_id = request.args.get('id', None, type=int)
    asset = Asset.query.get(asset_id) if asset_id else None
    
    if form.validate_on_submit():
        if asset:
            # 更新现有资产
            asset.ip = form.ip.data
            asset.name = form.name.data
        else:
            # 创建新资产
            asset = Asset(
                ip=form.ip.data,
                name=form.name.data,
                created_by=current_user.id
            )
            db.session.add(asset)
        db.session.commit()
        return redirect(url_for('assets'))
    
    if asset:
        form.ip.data = asset.ip
        form.name.data = asset.name
    
    assets = Asset.query.all()
    return render_template('assets.html', form=form, assets=assets, asset=asset)


@app.route('/delete_asset/<int:id>')
@login_required
def delete_asset(id):
    asset = Asset.query.get_or_404(id)
    db.session.delete(asset)
    db.session.commit()
    return redirect(url_for('assets'))


@app.route('/templates', methods=['GET', 'POST'])
@login_required
def templates():
    form = TemplateForm()
    template_id = request.args.get('id', None, type=int)
    template = ScanTemplate.query.get(template_id) if template_id else None
    
    if form.validate_on_submit():
        if template:
            # 更新现有模板
            template.name = form.name.data
            template.params = form.params.data
            template.description = form.description.data
        else:
            # 创建新模板
            template = ScanTemplate(
                name=form.name.data,
                params=form.params.data,
                description=form.description.data,
                created_by=current_user.id
            )
            db.session.add(template)
        db.session.commit()
        return redirect(url_for('templates'))
    
    if template:
        form.name.data = template.name
        form.params.data = template.params
        form.description.data = template.description
    
    templates = ScanTemplate.query.all()
    return render_template('templates.html', form=form, templates=templates, template=template)


@app.route('/delete_template/<int:id>', methods=['POST'])
@login_required
def delete_template(id):
    template = ScanTemplate.query.get_or_404(id)
    db.session.delete(template)
    db.session.commit()
    return redirect(url_for('templates'))


@app.route('/delete_task/<int:id>')
@login_required
def delete_task_old(id):
    task = Task.query.get_or_404(id)
    db.session.delete(task)
    db.session.commit()
    return redirect(url_for('tasks_list'))


@app.route('/users', methods=['GET', 'POST'])
@login_required
def manage_users():
    if not current_user.role == 'admin':
        return redirect(url_for('index'))
        
    form = UserForm()
    user_id = request.args.get('id', None, type=int)
    user = User.query.get(user_id) if user_id else None
    
    if form.validate_on_submit():
        if user:
            # 更新现有用户
            user.username = form.username.data
            user.role = form.role.data
            if form.password.data:
                user.password = generate_password_hash(form.password.data)
        else:
            # 创建新用户
            user = User(
                username=form.username.data,
                password=generate_password_hash(form.password.data),
                role=form.role.data
            )
            db.session.add(user)
        db.session.commit()
        return redirect(url_for('manage_users'))
    
    if user:
        form.username.data = user.username
        form.role.data = user.role
    
    users = User.query.all()
    return render_template('manage_users.html', form=form, users=users, user=user)


@app.route('/delete_user/<int:id>')
@login_required
def delete_user(id):
    if not current_user.role == 'admin':
        return redirect(url_for('index'))
        
    user = User.query.get_or_404(id)
    db.session.delete(user)
    db.session.commit()
    return redirect(url_for('manage_users'))


@app.route('/create-user', methods=['POST'])
@login_required
def create_user():
    if not current_user.role == 'admin':
        return jsonify({'success': False, 'error': '权限不足'}), 403
        
    form = UserForm()
    # 手动填充表单数据
    form.username.data = request.form.get('username')
    form.password.data = request.form.get('password')
    form.confirm.data = request.form.get('confirm')
    form.role.data = request.form.get('role')
    form.email.data = request.form.get('email')
    
    if form.validate():
        # 检查用户名是否已存在
        existing_user = User.query.filter_by(username=form.username.data).first()
        if existing_user:
            return jsonify({'success': False, 'error': '用户名已存在'})
        
        # 创建新用户
        user = User(
            username=form.username.data,
            password=generate_password_hash(form.password.data),
            role=form.role.data,
            email=form.email.data
        )
        db.session.add(user)
        db.session.commit()
        return jsonify({'success': True})
    
    # 表单验证失败，返回错误信息
    errors = {}
    for field, field_errors in form.errors.items():
        errors[field] = field_errors[0]  # 只取第一个错误信息
    
    return jsonify({'success': False, 'error': '表单验证失败', 'errors': errors})


@app.route('/edit-user/<int:user_id>', methods=['POST'])
@login_required
def edit_user(user_id):
    if not current_user.role == 'admin':
        return jsonify({'success': False, 'error': '权限不足'}), 403
        
    user = User.query.get_or_404(user_id)
    form = UserForm()
    
    # 手动填充表单数据
    form.username.data = request.form.get('username')
    form.password.data = request.form.get('password')
    form.confirm.data = request.form.get('confirm')
    form.role.data = request.form.get('role')
    form.email.data = request.form.get('email')
    
    if form.validate():
        user.username = form.username.data
        user.role = form.role.data
        user.email = form.email.data
        if form.password.data:  # 只有当提供了新密码时才更新密码
            user.password = generate_password_hash(form.password.data)
        db.session.commit()
        return jsonify({'success': True})
    
    # 表单验证失败，返回错误信息
    errors = {}
    for field, field_errors in form.errors.items():
        errors[field] = field_errors[0]  # 只取第一个错误信息
    
    return jsonify({'success': False, 'error': '表单验证失败', 'errors': errors})


@app.route('/email', methods=['GET', 'POST'])
@login_required
def email_settings():
    if not current_user.role == 'admin':
        return redirect(url_for('index'))
        
    form = EmailForm()
    email_setting = ZEmail.query.first()
    
    if form.validate_on_submit():
        if email_setting:
            # 更新现有设置
            email_setting.host = form.host.data
            email_setting.port = form.port.data
            email_setting.user = form.user.data
            email_setting.passwd = form.passwd.data
        else:
            # 创建新设置
            email_setting = ZEmail(
                host=form.host.data,
                port=form.port.data,
                user=form.user.data,
                passwd=form.passwd.data
            )
            db.session.add(email_setting)
        db.session.commit()
        return redirect(url_for('email_settings'))
    
    if email_setting:
        form.host.data = email_setting.host
        form.port.data = email_setting.port
        form.user.data = email_setting.user
        form.passwd.data = email_setting.passwd
    
    return render_template('email.html', form=form, email_setting=email_setting)


@app.route('/test_email', methods=['POST'])
@login_required
def test_email():
    if not current_user.role == 'admin':
        return jsonify({'success': False, 'error': '权限不足'})
    
    data = request.get_json()
    email_address = data.get('email')
    
    if not email_address:
        return jsonify({'success': False, 'error': '邮箱地址不能为空'})
    
    try:
        # 获取邮箱配置
        email_config = ZEmail.query.first()
        if not email_config:
            return jsonify({'success': False, 'error': '未配置邮箱信息'})
        
        # 创建测试邮件
        msg = MIMEMultipart()
        msg['From'] = email_config.user
        msg['To'] = email_address
        msg['Subject'] = Header('端口扫描系统测试邮件', 'utf-8')
        
        body = "这是一封测试邮件，用于验证邮箱配置是否正确。"
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # 发送邮件
        server = smtplib.SMTP(email_config.host, email_config.port)
        server.starttls()
        server.login(email_config.user, email_config.passwd)
        server.send_message(msg)
        server.quit()
        
        return jsonify({'success': True, 'message': '测试邮件发送成功'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/access-log')
@login_required
def access_log():
    if not current_user.role == 'admin':
        return redirect(url_for('index'))
        
    logs = AccessLog.query.order_by(AccessLog.login_time.desc()).all()
    return render_template('access_log.html', logs=logs)


@app.route('/get_user/<int:user_id>')
@login_required
def get_user(user_id):
    if not current_user.role == 'admin':
        return jsonify({'error': '权限不足'}), 403
        
    user = User.query.get_or_404(user_id)
    return jsonify({
        'id': user.id,
        'username': user.username,
        'role': user.role,
        'email': user.email
    })


@app.route('/run_task/<int:task_id>')
@login_required
def run_task(task_id):
    task = Task.query.get_or_404(task_id)
    # Run scan in background
    executor.submit(run_scan, task.id)
    flash('任务已重新启动，正在后台扫描', 'success')
    return redirect(url_for('tasks_list'))


@app.route('/execute_task/<int:task_id>', methods=['POST'])
def execute_task(task_id):
    """执行任务"""
    try:
        task = Task.query.get(task_id)
        if not task:
            return jsonify({'success': False, 'message': '任务不存在'})
        
        # 更新任务状态为运行中
        task.status = 'running'
        task.progress = 0
        db.session.commit()
        
        # 提交任务到后台执行
        executor.submit(run_scan, task_id)
        
        return jsonify({'success': True, 'message': '任务已开始执行'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


@app.route('/get_progress/<int:task_id>')
@login_required
def get_progress(task_id):
    """获取任务进度"""
    try:
        task = Task.query.get(task_id)
        if not task:
            return jsonify({'error': '任务不存在'}), 404
        
        # 检查用户权限（普通用户只能查看自己创建的任务，管理员可以查看所有任务）
        if session.get('role') != 'admin' and task.created_by != session['user_id']:
            return jsonify({'error': '权限不足'}), 403
        
        return jsonify({'progress': task.progress or 0})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/get_ips/<int:task_id>')
def get_ips(task_id):
    """获取任务的IP地址"""
    try:
        task = Task.query.get(task_id)
        if not task:
            return jsonify({'ips': []})
        
        ips = task.ips.split('\n')
        return jsonify({'ips': ips})
    except Exception as e:
        return jsonify({'ips': []})


@app.route('/delete_task/<int:task_id>', methods=['POST'])
def delete_task(task_id):
    """删除任务"""
    try:
        task = Task.query.get(task_id)
        if not task:
            return jsonify({'success': False, 'message': '任务不存在'})
        
        db.session.delete(task)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '任务删除成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


def run_scan(task_id):
    with app.app_context():
        task = db.session.get(Task, task_id)
        if not task:
            print(f"任务 {task_id} 不存在")
            return
            
        try:
            print(f"开始执行任务 {task_id}")
            task.status = 'running'
            task.progress = 0
            db.session.commit()
            print(f"任务 {task_id} 状态已更新为运行中")
            
            # 解析IP地址
            ip_list = []
            for line in task.ips.split('\n'):
                line = line.strip()
                if line:
                    ip_list.append(line)
            
            total_ips = len(ip_list)
            if total_ips == 0:
                print("未找到有效的IP地址")
                task.status = 'failed'
                task.finished_at = datetime.now()
                db.session.commit()
                return
                
            scan_results = []
            
            # 使用线程池并发执行扫描
            def scan_host(ip):
                """扫描单个主机"""
                if terminate_flags.get(task_id, False):
                    return None
                    
                try:
                    nm = nmap.PortScanner()
                    # 执行扫描
                    nm.scan(ip, arguments=task.params or '-p 1-1000')
                    
                    # 处理扫描结果
                    host_results = []
                    for host in nm.all_hosts():
                        host_data = {
                            'host': host,
                            'hostname': nm[host].hostname() if nm[host].hostname() else 'N/A',
                            'state': nm[host].state(),
                            'ports': []
                        }
                        
                        # 解决端口扫描结果存在close和filtered的问题，只保留open状态的端口
                        for protocol in nm[host].all_protocols():
                            ports = nm[host][protocol].keys()
                            for port in ports:
                                port_info = nm[host][protocol][port]
                                # 只添加开放的端口
                                if port_info.get('state') == 'open':
                                    port_data = {
                                        'port': port,
                                        'state': port_info.get('state', 'N/A'),
                                        'service': port_info.get('name', 'unknown'),
                                        'product': port_info.get('product', 'N/A'),
                                        'version': port_info.get('version', 'N/A'),
                                        'protocol': protocol
                                    }
                                    host_data['ports'].append(port_data)

                        host_results.append(host_data)
                    return host_results
                except Exception as e:
                    print(f"扫描IP {ip} 时出错: {str(e)}")
                    return None
            
            # 并发执行所有扫描任务
            print(f"开始扫描 {total_ips} 个IP地址")
            futures = []
            for ip in ip_list:
                future = executor.submit(scan_host, ip)
                futures.append((future, ip))
            
            completed = 0
            for future, ip in futures:
                if terminate_flags.get(task_id, False):
                    task.status = 'cancelled'
                    task.finished_at = datetime.now()
                    db.session.commit()
                    return
                    
                try:
                    result = future.result(timeout=30)  # 30秒超时
                    if result:
                        scan_results.extend(result)
                except concurrent.futures.TimeoutError:
                    print(f"扫描IP {ip} 超时")
                except Exception as e:
                    print(f"扫描IP {ip} 出错: {str(e)}")
                
                completed += 1
                task.progress = int((completed / total_ips) * 100)
                db.session.commit()
                print(f"已完成 {completed}/{total_ips} 个IP的扫描, 进度: {task.progress}%")
            
            # 保存结果
            formatted_results = []
            for host_data in scan_results:
                formatted_host = {
                    'host': host_data['host'],
                    'hostname': host_data['hostname'],
                    'state': host_data['state'],
                    'ports': []
                }

                for port_data in host_data['ports']:
                    # 只添加开放的端口
                    if port_data['state'] == 'open':
                        formatted_port = {
                            'port': port_data['port'],
                            'state': port_data['state'],
                            'service': port_data['service'],
                            'product': port_data['product'],
                            'version': port_data['version'],
                            'protocol': port_data['protocol']
                        }
                        formatted_host['ports'].append(formatted_port)

                formatted_results.append(formatted_host)
            
            task.result = json.dumps(formatted_results, ensure_ascii=False)
            task.status = 'completed'
            task.progress = 100
            task.finished_at = datetime.now()
            db.session.commit()
            print(f"任务 {task_id} 执行完成, 最终进度: {task.progress}%")
            
            # 如果需要发送邮件且配置了邮箱
            if task.send_email and task.email_address:
                try:
                    send_scan_result_email(task)
                except Exception as e:
                    print(f"发送邮件失败: {str(e)}")
                    
        except Exception as e:
            print(f"扫描任务 {task_id} 执行失败: {str(e)}")
            import traceback
            print(traceback.format_exc())
            task.status = 'failed'
            task.finished_at = datetime.now()
            db.session.commit()
        finally:
            # 清理终止标志
            if task_id in terminate_flags:
                del terminate_flags[task_id]


def send_scan_result_email(task):
    """发送扫描结果邮件"""
    # 获取邮箱配置
    email_config = ZEmail.query.first()
    if not email_config:
        raise Exception("未配置邮箱信息")
    
    # 解析扫描结果以生成统计数据
    total_hosts = 0
    total_open_ports = 0
    unique_services = set()
    
    if task.result:
        try:
            results = json.loads(task.result)
            total_hosts = len(results)
            
            for result in results:
                if 'ports' in result:
                    open_ports = [port for port in result['ports'] if port['state'] == 'open']
                    total_open_ports += len(open_ports)
                    
                    for port in open_ports:
                        if port['service'] != 'unknown':
                            unique_services.add(port['service'])
        except Exception as e:
            print(f"解析扫描结果时出错: {str(e)}")
    
    # 创建邮件
    msg = MIMEMultipart()
    msg['From'] = email_config.user
    msg['To'] = task.email_address
    msg['Subject'] = Header(f'端口扫描结果 - {task.name}', 'utf-8')
    
    # 构建邮件正文（包含分析报告摘要和统计数据）
    body = f"""
扫描任务已完成！

任务名称: {task.name}
扫描时间: {task.finished_at.strftime('%Y-%m-%d %H:%M:%S')}

扫描结果统计:
  - 扫描主机数量: {total_hosts}
  - 开放端口总数: {total_open_ports}
  - 发现服务种类: {len(unique_services)}

这是端口扫描任务的分析报告摘要。完整的扫描结果请查看附件中的详细报告。
    
附件中包含完整的扫描结果，您可以下载并查看详细的端口信息。
"""
    
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    
    # 添加扫描结果作为附件(XLSX格式)
    if task.result:
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "扫描结果"
        
        # 设置标题行样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入表头
        headers = ['主机IP', '主机名', '状态', '端口', '协议', '服务', '产品', '版本']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
        
        # 写入数据
        row_num = 2
        results = json.loads(task.result)
        for result in results:
            host = result.get('host', 'N/A')
            hostname = result.get('hostname', 'N/A')
            state = result.get('state', 'N/A')
            
            if 'ports' in result and result['ports']:
                for port in result['ports']:
                    ws.cell(row=row_num, column=1, value=host)
                    ws.cell(row=row_num, column=2, value=hostname)
                    ws.cell(row=row_num, column=3, value=state)
                    ws.cell(row=row_num, column=4, value=port.get('port', 'N/A'))
                    ws.cell(row=row_num, column=5, value=port.get('protocol', 'N/A'))
                    ws.cell(row=row_num, column=6, value=port.get('service', 'unknown'))
                    ws.cell(row=row_num, column=7, value=port.get('product', 'N/A'))
                    ws.cell(row=row_num, column=8, value=port.get('version', 'N/A'))
                    row_num += 1
            else:
                # 没有端口信息的主机
                ws.cell(row=row_num, column=1, value=host)
                ws.cell(row=row_num, column=2, value=hostname)
                ws.cell(row=row_num, column=3, value=state)
                row_num += 1
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 将Excel文件保存到内存中
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # 创建附件
        attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        attachment.set_payload(excel_buffer.read())
        encoders.encode_base64(attachment)
        attachment.add_header(
            'Content-Disposition',
            'attachment',
            filename=f'scan_result_{task.name}_{task.id}.xlsx'
        )
        msg.attach(attachment)
    else:
        # 添加空结果提示附件
        empty_wb = Workbook()
        empty_ws = empty_wb.active
        empty_ws.title = "扫描结果"
        empty_ws.cell(row=1, column=1, value="未发现任何扫描结果")
        
        empty_excel_buffer = io.BytesIO()
        empty_wb.save(empty_excel_buffer)
        empty_excel_buffer.seek(0)
        
        attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        attachment.set_payload(empty_excel_buffer.read())
        encoders.encode_base64(attachment)
        attachment.add_header(
            'Content-Disposition',
            'attachment',
            filename='scan_result_empty.xlsx'
        )
        msg.attach(attachment)
    
    # 发送邮件
    try:
        server = smtplib.SMTP(email_config.host, email_config.port)
        server.starttls()
        server.login(email_config.user, email_config.passwd)
        server.send_message(msg)
        server.quit()
    except Exception as e:
        raise Exception(f"发送邮件失败: {str(e)}")

@app.route('/fofa')
@login_required
def fofa():
    """FOFA资产搜索页面"""
    return render_template('fofa.html')

@app.route('/api/fofa/search', methods=['POST'])
@login_required
def fofa_search():
    """FOFA搜索API"""
    try:
        data = request.get_json()
        query = data.get('query', '')
        size = data.get('size', 100)
        
        if not query:
            return jsonify({'success': False, 'error': '查询语句不能为空'}), 400
        
        # 限制最大返回数量
        size = min(size, 10000)
        
        # 这里应该实现真实的FOFA API调用
        # 目前返回模拟数据用于演示
        results = []
        for i in range(min(size, 100)):  # 限制模拟数据数量
            results.append({
                'ip': f'192.168.1.{i+1}',
                'port': 80,
                'protocol': 'http',
                'domain': f'example{i+1}.com',
                'title': f'Example Site {i+1}',
                'server': 'Apache' if i % 2 == 0 else 'Nginx',
                'os': 'Linux' if i % 3 == 0 else 'Windows',
                'country': 'CN'
            })
        
        return jsonify({
            'success': True,
            'results': results,
            'total': len(results)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/asset_center')
@login_required
def asset_center():
    """资产中心页面"""
    # 获取所有已完成的扫描任务
    tasks = Task.query.filter_by(status='completed').all()
    return render_template('asset_center.html', tasks=tasks)

@app.route('/api/assets/data')
@login_required
def api_assets_data():
    """获取资产数据API"""
    try:
        # 获取所有已完成的扫描任务
        tasks = Task.query.filter_by(status='completed').all()
        
        assets = []
        for task in tasks:
            if task.result:
                try:
                    result_data = json.loads(task.result)
                    for host_data in result_data:
                        host = host_data.get('host', 'N/A')
                        if 'ports' in host_data and isinstance(host_data['ports'], list):
                            for port_data in host_data['ports']:
                                asset = {
                                    'ip': host,
                                    'port': port_data.get('port', 'N/A'),
                                    'protocol': port_data.get('protocol', 'tcp'),
                                    'service': port_data.get('service', 'unknown'),
                                    'product': port_data.get('product', ''),
                                    'version': port_data.get('version', ''),
                                    'task_id': task.id,
                                    'task_name': task.name,
                                    'scan_time': task.finished_at.strftime('%Y-%m-%d %H:%M:%S') if task.finished_at else 'N/A'
                                }
                                assets.append(asset)
                except json.JSONDecodeError:
                    # 忽略无法解析的结果
                    continue
        
        return jsonify({
            'success': True,
            'assets': assets
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/terminate_task/<int:task_id>', methods=['POST'])
@login_required
def terminate_task(task_id):
    task = Task.query.get(task_id)
    if not task:
        return jsonify({'success': False, 'error': '任务不存在'}), 404
    
    # 检查用户权限（普通用户只能终止自己创建的任务，管理员可以终止所有任务）
    if session.get('role') != 'admin' and task.created_by != session['user_id']:
        return jsonify({'success': False, 'error': '权限不足'}), 403
    
    # 设置终止标志
    terminate_flags[task_id] = True
    
    # 更新任务状态
    task.status = 'terminated'
    db.session.commit()
    
    return jsonify({'success': True})


def scan_udp_ports(ip, params, task_id):
    """扫描UDP端口"""
    if terminate_flags.get(task_id, False):
        return None
            
    try:
        nm = nmap.PortScanner()
        nm.scan(ip, arguments=params)
            
        results = []
        for host in nm.all_hosts():
            host_data = {
                'host': host,
                'hostname': nm[host].hostname() if nm[host].hostname() else 'N/A',
                'state': nm[host].state(),
                'ports': []
            }
                
            for protocol in nm[host].all_protocols():
                if protocol in nm[host]:
                    ports = nm[host][protocol].keys()
                    for port in ports:
                        port_info = nm[host][protocol][port]
                        port_data = {
                            'port': port,
                            'state': port_info.get('state', 'N/A'),
                            'service': port_info.get('name', 'unknown'),
                            'product': port_info.get('product', 'N/A'),
                            'version': port_info.get('version', 'N/A'),
                            'protocol': protocol
                        }
                        host_data['ports'].append(port_data)

            for protocol in nm[host].all_protocols():
                ports = nm[host][protocol].keys()
                for port in ports:
                    port_info = nm[host][protocol][port]
                    # 只添加开放的端口
                    if port_info.get('state') == 'open':
                        port_data = {
                            'port': port,
                            'state': port_info.get('state', 'N/A'),
                            'service': port_info.get('name', 'unknown'),
                            'product': port_info.get('product', 'N/A'),
                            'version': port_info.get('version', 'N/A'),
                            'protocol': protocol
                        }
                        host_data['ports'].append(port_data)

                results.append(host_data)
            
        return results
    except Exception as e:
        print(f"扫描UDP端口时出错: {str(e)}")
        return None

def execute_scan_task(task_id):
    """执行扫描任务"""
    task = ScanTask.query.get(task_id)
    if not task:
        return
            
    try:
        # 获取IP列表
        ip_list = []
        if task.asset_id:
            asset = Asset.query.get(task.asset_id)
            if asset:
                ip_list = [ip.strip() for ip in asset.ip.split('\n') if ip.strip()]
            
        if task.custom_ips:
            custom_ips = [ip.strip() for ip in task.custom_ips.split('\n') if ip.strip()]
            ip_list.extend(custom_ips)
            
        # 去重
        ip_list = list(set(ip_list))
        total_ips = len(ip_list)
            
        if total_ips == 0:
            print("未找到有效的IP地址")
            task.status = 'failed'
            task.finished_at = datetime.now()
            db.session.commit()
            return
                
        # 更新任务状态
        task.status = 'running'
        task.started_at = datetime.now()
        task.progress = 0
        db.session.commit()
            
        scan_results = []
            
        # 使用线程池并发执行扫描
        def scan_host(ip):
            """扫描单个主机"""
            if terminate_flags.get(task_id, False):
                return None
                    
            try:
                nm = nmap.PortScanner()
                # 执行扫描
                nm.scan(ip, arguments=task.params or '-p 1-1000')
                    
                # 处理扫描结果
                host_results = []
                for host in nm.all_hosts():
                    host_data = {
                        'host': host,
                        'hostname': nm[host].hostname() if nm[host].hostname() else 'N/A',
                        'state': nm[host].state(),
                        'ports': []
                    }
                        
                    for protocol in nm[host].all_protocols():
                        ports = nm[host][protocol].keys()
                        for port in ports:
                            port_info = nm[host][protocol][port]
                            port_data = {
                                'port': port,
                                'state': port_info.get('state', 'N/A'),
                                'service': port_info.get('name', 'unknown'),
                                'product': port_info.get('product', 'N/A'),
                                'version': port_info.get('version', 'N/A'),
                                'protocol': protocol
                            }
                            host_data['ports'].append(port_data)
                        
                    host_results.append(host_data)
                return host_results
            except Exception as e:
                print(f"扫描IP {ip} 时出错: {str(e)}")
                return None
            
        # 并发执行所有扫描任务
        print(f"开始扫描 {total_ips} 个IP地址")
        futures = []
        for ip in ip_list:
            future = executor.submit(scan_host, ip)
            futures.append((future, ip))
            
        completed = 0
        for future, ip in futures:
            if terminate_flags.get(task_id, False):
                task.status = 'cancelled'
                task.finished_at = datetime.now()
                db.session.commit()
                return
                    
            try:
                result = future.result(timeout=30)  # 30秒超时
                if result:
                    scan_results.extend(result)
            except concurrent.futures.TimeoutError:
                print(f"扫描IP {ip} 超时")
            except Exception as e:
                print(f"扫描IP {ip} 出错: {str(e)}")
                
            completed += 1
            task.progress = int((completed / total_ips) * 100)
            db.session.commit()
            print(f"已完成 {completed}/{total_ips} 个IP的扫描, 进度: {task.progress}%")
            
        # 保存结果
        formatted_results = []
        for host_data in scan_results:
            formatted_host = {
                'host': host_data['host'],
                'hostname': host_data['hostname'],
                'state': host_data['state'],
                'ports': []
            }
                

            for port_data in host_data['ports']:
                # 只添加开放的端口
                if port_data['state'] == 'open':
                    formatted_port = {
                        'port': port_data['port'],
                        'state': port_data['state'],
                        'service': port_data['service'],
                        'product': port_data['product'],
                        'version': port_data['version'],
                        'protocol': port_data['protocol']
                    }
                    formatted_host['ports'].append(formatted_port)             
            formatted_results.append(formatted_host)
            
        task.result = json.dumps(formatted_results, ensure_ascii=False)
        task.status = 'completed'
        task.progress = 100
        task.finished_at = datetime.now()
        db.session.commit()
            
    except Exception as e:
        print(f"执行扫描任务时出错: {str(e)}")
        task.status = 'failed'
        task.finished_at = datetime.now()
        db.session.commit()
    finally:
        # 清理终止标志
        if task_id in terminate_flags:
            del terminate_flags[task_id]

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        # 创建默认管理员账户
        admin = User.query.filter_by(username='admin').first()
        if not admin:
            admin_user = User(
                username='admin',
                password=generate_password_hash('admin'),
                role='admin',
                email='admin@example.com'
            )
            db.session.add(admin_user)
            db.session.commit()
            print("默认管理员账户已创建: admin/admin")
    app.run(debug=True, host='0.0.0.0')